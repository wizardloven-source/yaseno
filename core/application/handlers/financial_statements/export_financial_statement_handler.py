# core/application/handlers/financial_statements/export_financial_statement_handler.py
"""
Export Financial Statement Handler - معالج تصدير القوائم المالية
"""

import logging
import json
from datetime import datetime
from typing import Dict, Any, Optional
from pathlib import Path

from core.domain.financial_statements.value_objects import StatementId
from core.domain.accounting.interfaces import IUnitOfWork

from core.application.handlers.base_handler import BaseHandler
from core.application.security.authorization import UserContext, require_permission, Permission
from core.application.financial_statements.commands import ExportFinancialStatementCommand
from core.application.financial_statements.dtos import FinancialStatementDTO
from core.application.financial_statements.converters import statement_to_dict

logger = logging.getLogger(__name__)


class ExportFinancialStatementHandler(BaseHandler[ExportFinancialStatementCommand, Dict[str, Any]]):
    """
    معالج تصدير القوائم المالية
    
    يقوم بتصدير القائمة المالية إلى صيغ مختلفة (JSON, CSV, Excel, PDF)
    """

    def __init__(self, uow: IUnitOfWork):
        super().__init__(uow)

    @require_permission(Permission.VIEW_TRIAL_BALANCE)
    def handle(self, command: ExportFinancialStatementCommand, user_context: UserContext) -> Dict[str, Any]:
        """
        تنفيذ تصدير القائمة المالية
        
        Args:
            command: أمر تصدير القائمة المالية
            user_context: سياق المستخدم
        
        Returns:
            Dict[str, Any]: نتيجة التصدير
        """
        logger.info(f"Exporting financial statement: {command.statement_id} as {command.format}")

        with self._uow:
            # جلب القائمة المالية
            statement = self._uow.financial_statements.get_by_id(
                StatementId(command.statement_id)
            )

            if not statement:
                return {
                    "success": False,
                    "message": f"Financial statement {command.statement_id} not found",
                    "statement_id": command.statement_id
                }

            # تحويل القائمة إلى قاموس
            statement_data = statement_to_dict(statement)

            # تحديد مسار التصدير
            export_path = command.export_path or self._get_default_export_path(statement)

            # تصدير حسب الصيغة المطلوبة
            if command.format == "json":
                return self._export_json(statement_data, export_path, user_context)
            elif command.format == "csv":
                return self._export_csv(statement_data, export_path, user_context)
            elif command.format == "excel":
                return self._export_excel(statement_data, export_path, user_context)
            elif command.format == "pdf":
                return self._export_pdf(statement_data, export_path, user_context)
            else:
                return {
                    "success": False,
                    "message": f"Unsupported export format: {command.format}",
                    "statement_id": command.statement_id
                }

    def _get_default_export_path(self, statement) -> str:
        """الحصول على مسار التصدير الافتراضي"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        statement_type = statement.statement_type.value if hasattr(statement.statement_type, 'value') else str(statement.statement_type)
        return f"exports/{statement_type}_{timestamp}"

    def _export_json(self, data: Dict, path: str, user_context: UserContext) -> Dict[str, Any]:
        """تصدير إلى JSON"""
        try:
            file_path = f"{path}.json"
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Financial statement exported to JSON: {file_path}")
            
            return {
                "success": True,
                "message": f"Statement exported successfully to {file_path}",
                "file_path": file_path,
                "format": "json",
                "statement_id": data.get('id'),
                "exported_by": user_context.user_id,
                "exported_at": datetime.now().isoformat()
            }
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            return {
                "success": False,
                "message": f"Error exporting to JSON: {str(e)}",
                "statement_id": data.get('id')
            }

    def _export_csv(self, data: Dict, path: str, user_context: UserContext) -> Dict[str, Any]:
        """تصدير إلى CSV"""
        try:
            import csv
            file_path = f"{path}.csv"
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # استخراج البيانات من الأقسام
            rows = []
            for section in data.get('sections', []):
                for line in section.get('lines', []):
                    rows.append({
                        'section': section.get('name', ''),
                        'account_code': line.get('code', ''),
                        'account_name': line.get('name', ''),
                        'amount': line.get('amount', 0),
                        'currency': line.get('currency', 'USD'),
                        'level': line.get('level', 0),
                        'is_total': line.get('is_total', False),
                        'is_subtotal': line.get('is_subtotal', False)
                    })
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                if rows:
                    writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                    writer.writeheader()
                    writer.writerows(rows)
            
            logger.info(f"Financial statement exported to CSV: {file_path}")
            
            return {
                "success": True,
                "message": f"Statement exported successfully to {file_path}",
                "file_path": file_path,
                "format": "csv",
                "rows_exported": len(rows),
                "statement_id": data.get('id'),
                "exported_by": user_context.user_id,
                "exported_at": datetime.now().isoformat()
            }
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return {
                "success": False,
                "message": f"Error exporting to CSV: {str(e)}",
                "statement_id": data.get('id')
            }

    def _export_excel(self, data: Dict, path: str, user_context: UserContext) -> Dict[str, Any]:
        """تصدير إلى Excel"""
        try:
            try:
                import openpyxl
                from openpyxl import Workbook
            except ImportError:
                return {
                    "success": False,
                    "message": "openpyxl is not installed. Please install it: pip install openpyxl",
                    "statement_id": data.get('id')
                }
            
            file_path = f"{path}.xlsx"
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Statement"
            
            # إضافة العنوان
            ws['A1'] = f"Financial Statement: {data.get('type', '')}"
            ws['A2'] = f"Period: {data.get('period_start', '')} to {data.get('period_end', '')}"
            ws['A3'] = f"Currency: {data.get('currency', 'USD')}"
            ws['A4'] = f"Generated at: {data.get('generated_at', '')}"
            
            # إضافة الأقسام والأسطر
            row = 6
            for section in data.get('sections', []):
                ws.cell(row=row, column=1, value=section.get('name', ''))
                ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
                row += 1
                
                for line in section.get('lines', []):
                    ws.cell(row=row, column=1, value=line.get('code', ''))
                    ws.cell(row=row, column=2, value=line.get('name', ''))
                    ws.cell(row=row, column=3, value=line.get('amount', 0))
                    ws.cell(row=row, column=4, value=line.get('currency', 'USD'))
                    row += 1
                
                # إضافة المجموع الفرعي
                ws.cell(row=row, column=2, value=f"Total {section.get('name', '')}")
                ws.cell(row=row, column=3, value=section.get('total', 0))
                ws.cell(row=row, column=3).font = openpyxl.styles.Font(bold=True)
                row += 2
            
            # حفظ الملف
            wb.save(file_path)
            
            logger.info(f"Financial statement exported to Excel: {file_path}")
            
            return {
                "success": True,
                "message": f"Statement exported successfully to {file_path}",
                "file_path": file_path,
                "format": "excel",
                "statement_id": data.get('id'),
                "exported_by": user_context.user_id,
                "exported_at": datetime.now().isoformat()
            }
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return {
                "success": False,
                "message": f"Error exporting to Excel: {str(e)}",
                "statement_id": data.get('id')
            }

    def _export_pdf(self, data: Dict, path: str, user_context: UserContext) -> Dict[str, Any]:
        """تصدير إلى PDF"""
        try:
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
            except ImportError:
                return {
                    "success": False,
                    "message": "reportlab is not installed. Please install it: pip install reportlab",
                    "statement_id": data.get('id')
                }
            
            file_path = f"{path}.pdf"
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # إضافة العنوان
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1  # Center
            )
            story.append(Paragraph(f"Financial Statement: {data.get('type', '')}", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # إضافة معلومات الفترة
            info_style = ParagraphStyle(
                'Info',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#7f8c8d')
            )
            story.append(Paragraph(f"Period: {data.get('period_start', '')} to {data.get('period_end', '')}", info_style))
            story.append(Paragraph(f"Currency: {data.get('currency', 'USD')}", info_style))
            story.append(Paragraph(f"Generated: {data.get('generated_at', '')}", info_style))
            story.append(Spacer(1, 0.3*inch))
            
            # إضافة الجداول
            for section in data.get('sections', []):
                # عنوان القسم
                section_style = ParagraphStyle(
                    'Section',
                    parent=styles['Heading2'],
                    fontSize=12,
                    textColor=colors.HexColor('#34495e')
                )
                story.append(Paragraph(section.get('name', ''), section_style))
                story.append(Spacer(1, 0.1*inch))
                
                # جدول الأسطر
                table_data = [['Code', 'Name', 'Amount', 'Currency']]
                for line in section.get('lines', []):
                    table_data.append([
                        line.get('code', ''),
                        line.get('name', ''),
                        f"{line.get('amount', 0):,.2f}",
                        line.get('currency', 'USD')
                    ])
                
                if len(table_data) > 1:
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    story.append(table)
                story.append(Spacer(1, 0.2*inch))
            
            # بناء المستند
            doc.build(story)
            
            logger.info(f"Financial statement exported to PDF: {file_path}")
            
            return {
                "success": True,
                "message": f"Statement exported successfully to {file_path}",
                "file_path": file_path,
                "format": "pdf",
                "statement_id": data.get('id'),
                "exported_by": user_context.user_id,
                "exported_at": datetime.now().isoformat()
            }
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            return {
                "success": False,
                "message": f"Error exporting to PDF: {str(e)}",
                "statement_id": data.get('id')
            }